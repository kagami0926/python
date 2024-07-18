import openpyxl
import matplotlib.pyplot as plt
import japanize_matplotlib
import math
import numpy as np

# Excelファイルからデータを読み込む
wb = openpyxl.load_workbook("3.law of conservation of momentum/data/com.xlsx")
ws1 = wb["Sheet1"]

# データの集合countとxを設定
count = []
x = []

# データを取得する
for row in ws1['H4:H19']:
    for cell in row:
        count.append(cell.value)

for row in ws1['I4:I19']:
    for cell in row:
        x.append(cell.value)

# phiとlogphiを計算する
l = 78
phi = []
logphi = []

for h in x:
    if h is not None:
        phi_value = math.sqrt((2 * l - 2 * math.sqrt(l**2 - h**2)) / l)
        logphi_value = math.log(phi_value)
        phi.append(phi_value)
        logphi.append(logphi_value)
    else:
        phi.append(None)
        logphi.append(None)

# phiとlogphiをExcelファイルに書き込む
start_row = 4  # 書き込みを開始する行（4行目から）

for i in range(len(phi)):
    ws1.cell(row=start_row + i, column=11, value=phi[i])
    ws1.cell(row=start_row + i, column=12, value=logphi[i])

# Excelファイルに保存する
wb.save("3.law of conservation of momentum/data/com.xlsx")

# ワークブックを閉じる
wb.close()

# Noneを除いたフィルタリングされたデータを作成
filtered_count = []
filtered_logphi = []

for i in range(len(phi)):
    if phi[i] is not None:
        filtered_count.append(count[i])
        filtered_logphi.append(logphi[i])

# 散布図を作成
plt.scatter(filtered_count, filtered_logphi)

# フィッティングを実行
coefficients = np.polyfit(filtered_count, filtered_logphi, 1)  # 1は一次式（直線）を意味する
slope, intercept = coefficients

# フィットした直線の方程式
fit_line = slope * np.array(filtered_count) + intercept

# フィットした直線をプロット
plt.plot(filtered_count, fit_line, color='red', label=f'近似直線: y = {slope:.2f}x + {intercept:.2f}')

# グラフの装飾
plt.title('小球の衝突回数と最大振れ角の関係')
plt.xlabel('小球の衝突回数n(回)')
plt.ylabel(r"$\log \phi_n$")
plt.grid()
plt.legend()

# グラフを画像ファイルとして保存
plt.savefig("3.law of conservation of momentum\images\com.png")

# グラフを表示
plt.show()

print(f"Slope: {slope}")
print(f"Intercept: {intercept}")
